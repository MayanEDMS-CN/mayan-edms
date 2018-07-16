import logging

from django.apps import apps
from django.contrib.auth import get_user_model
from django.core.files import File
from django.db import OperationalError
from django.utils.encoding import force_text
from django.utils.translation import ugettext_lazy as _
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook

from mayan.celery import app

from common.compressed_files import CompressedFile, NotACompressedFile
from documents.tasks import task_upload_new_version

from .literals import DEFAULT_SOURCE_TASK_RETRY_DELAY

logger = logging.getLogger(__name__)


def is_chery_excel_package(file_object, label=None):
    if label is None:
        return False
    if not label.strip().lower().endswith(".xlsx"):
        return False
    result = False
    try:
        wb = load_workbook(file_object, read_only=True)
        ws = wb.active
        result = '序号,话术类别,二级分类,问题,内容' == ",".join(cell.value.strip() for cell in ws.rows.__next__())
        wb.close()
    except:
        pass
    return result


@app.task(ignore_result=True)
def task_check_interval_source(source_id):
    Source = apps.get_model(
        app_label='sources', model_name='Source'
    )

    source = Source.objects.get_subclass(pk=source_id)
    if source.enabled:
        try:
            source.check_source()
        except Exception as exception:
            logger.error('Error processing source: %s; %s', source, exception)
            source.logs.create(
                message=_('Error processing source: %s') % exception
            )
        else:
            source.logs.all().delete()


@app.task(bind=True, default_retry_delay=DEFAULT_SOURCE_TASK_RETRY_DELAY, ignore_result=True)
def task_upload_document(self, source_id, document_type_id, shared_uploaded_file_id, description=None, label=None, language=None, metadata_dict_list=None, tag_ids=None, user_id=None):
    SharedUploadedFile = apps.get_model(
        app_label='common', model_name='SharedUploadedFile'
    )

    DocumentType = apps.get_model(
        app_label='documents', model_name='DocumentType'
    )

    Source = apps.get_model(
        app_label='sources', model_name='Source'
    )

    try:
        document_type = DocumentType.objects.get(pk=document_type_id)
        source = Source.objects.get_subclass(pk=source_id)
        shared_upload = SharedUploadedFile.objects.get(
            pk=shared_uploaded_file_id
        )

        if user_id:
            user = get_user_model().objects.get(pk=user_id)
        else:
            user = None

        with shared_upload.open() as file_object:
            source.upload_document(
                file_object=file_object, document_type=document_type,
                description=description, label=label, language=language,
                metadata_dict_list=metadata_dict_list, user=user,
                tag_ids=tag_ids
            )

    except OperationalError as exception:
        logger.warning(
            'Operational exception while trying to create new document "%s" '
            'from source id %d; %s. Retying.',
            label or shared_upload.filename, source_id, exception
        )
        raise self.retry(exc=exception)
    else:
        try:
            shared_upload.delete()
        except OperationalError as exception:
            logger.warning(
                'Operational error during attempt to delete shared upload '
                'file: %s; %s. Retrying.', shared_upload, exception
            )


@app.task(bind=True, default_retry_delay=DEFAULT_SOURCE_TASK_RETRY_DELAY, ignore_result=True)
def task_source_handle_upload(self, document_type_id, shared_uploaded_file_id, source_id, description=None, expand=False, label=None, language=None, metadata_dict_list=None, skip_list=None, tag_ids=None, user_id=None):
    SharedUploadedFile = apps.get_model(
        app_label='common', model_name='SharedUploadedFile'
    )

    DocumentType = apps.get_model(
        app_label='documents', model_name='DocumentType'
    )

    Document = apps.get_model(
        app_label='documents', model_name='Document'
    )

    try:
        document_type = DocumentType.objects.get(pk=document_type_id)
        shared_upload = SharedUploadedFile.objects.get(
            pk=shared_uploaded_file_id
        )

        if not label:
            label = shared_upload.filename

    except OperationalError as exception:
        logger.warning(
            'Operational error during attempt to load data to handle source '
            'upload: %s. Retrying.', exception
        )
        raise self.retry(exc=exception)

    kwargs = {
        'description': description, 'document_type_id': document_type.pk,
        'label': label, 'language': language,
        'metadata_dict_list': metadata_dict_list,
        'source_id': source_id, 'tag_ids': tag_ids, 'user_id': user_id
    }

    if not skip_list:
        skip_list = []

    with shared_upload.open() as file_object:
        if expand:
            try:
                compressed_file = CompressedFile(file_object)
                for compressed_file_child in compressed_file.children():
                    # TODO: find way to uniquely identify child files
                    # Use filename in the mean time.
                    if force_text(compressed_file_child) not in skip_list:
                        kwargs.update(
                            {'label': force_text(compressed_file_child)}
                        )

                        try:
                            child_shared_uploaded_file = SharedUploadedFile.objects.create(
                                file=File(compressed_file_child)
                            )
                        except OperationalError as exception:
                            logger.warning(
                                'Operational error while preparing to upload '
                                'child document: %s. Rescheduling.', exception
                            )

                            task_source_handle_upload.delay(
                                document_type_id=document_type_id,
                                shared_uploaded_file_id=shared_uploaded_file_id,
                                source_id=source_id, description=description,
                                expand=expand, label=label,
                                language=language,
                                metadata_dict_list=metadata_dict_list,
                                skip_list=skip_list, tag_ids=tag_ids,
                                user_id=user_id
                            )
                            return
                        else:
                            skip_list.append(force_text(compressed_file_child))
                            task_upload_document.delay(
                                shared_uploaded_file_id=child_shared_uploaded_file.pk,
                                **kwargs
                            )
                        finally:
                            compressed_file_child.close()

                    compressed_file_child.close()
                try:
                    shared_upload.delete()
                except OperationalError as exception:
                    logger.warning(
                        'Operational error during attempt to delete shared '
                        'upload file: %s; %s. Retrying.', shared_upload,
                        exception
                    )
            except NotACompressedFile:
                logging.debug('Exception: NotACompressedFile')
                task_upload_document.delay(
                    shared_uploaded_file_id=shared_upload.pk, **kwargs
                )
        elif is_chery_excel_package(shared_upload.open(), label):
            wb = load_workbook(shared_upload.open(), read_only=True)
            ws = wb.active
            for id, cat1, cat2, title, content in ws.rows:
                filename = title.value
                if filename == '问题':
                    continue
                file_child = SimpleUploadedFile(
                    name=filename, content=content.value.encode("utf8")
                )
                child_shared_uploaded_file = SharedUploadedFile.objects.create(
                    file=File(file_child)
                )
                possible_duplicated = Document.objects.filter(label=filename).last()
                if possible_duplicated is None:
                    kwargs.update({
                        "label": filename,
                        "description": filename
                    })
                    task_upload_document.delay(
                        shared_uploaded_file_id=child_shared_uploaded_file.pk, **kwargs
                    )
                else:
                    task_upload_new_version.delay(
                        document_id = possible_duplicated.id,
                        shared_uploaded_file_id=child_shared_uploaded_file.pk,
                        user_id=user_id
                    )
        else:
            task_upload_document.delay(
                shared_uploaded_file_id=shared_upload.pk, **kwargs
            )
